"""
odds_backfill.py — Historical odds backfill for NFL and MLB.

Downloads free Excel datasets, maps team names to ESPN display names,
joins to existing fact_game_results rows by (game_date, home_team, away_team),
and upserts into fact_market_odds.

Data sources:
  - NFL: aussportsbetting.com (2006-present, single .xlsx)
  - MLB: sports-statistics.com (2010-2021, per-year .xlsx)

Public entry point: ``backfill_historical_odds(sport, start_year, end_year)``
"""

import io
import logging
from datetime import datetime

import requests

from database import get_connection

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Data source URLs
# ---------------------------------------------------------------------------

NFL_URL = "https://www.aussportsbetting.com/historical_data/nfl.xlsx"
MLB_URL_TEMPLATE = "https://sports-statistics.com/database/mlb-data/mlb-odds-{year}.xlsx"
MLB_YEARS = range(2010, 2022)  # 2010-2021 inclusive

# ---------------------------------------------------------------------------
# Team name aliases → ESPN display name
# ---------------------------------------------------------------------------
# Keys are lowercased source names. Values are ESPN displayName.

NFL_TEAM_ALIASES = {
    # Standard short names
    "arizona cardinals": "Arizona Cardinals",
    "atlanta falcons": "Atlanta Falcons",
    "baltimore ravens": "Baltimore Ravens",
    "buffalo bills": "Buffalo Bills",
    "carolina panthers": "Carolina Panthers",
    "chicago bears": "Chicago Bears",
    "cincinnati bengals": "Cincinnati Bengals",
    "cleveland browns": "Cleveland Browns",
    "dallas cowboys": "Dallas Cowboys",
    "denver broncos": "Denver Broncos",
    "detroit lions": "Detroit Lions",
    "green bay packers": "Green Bay Packers",
    "houston texans": "Houston Texans",
    "indianapolis colts": "Indianapolis Colts",
    "jacksonville jaguars": "Jacksonville Jaguars",
    "kansas city chiefs": "Kansas City Chiefs",
    "las vegas raiders": "Las Vegas Raiders",
    "los angeles chargers": "Los Angeles Chargers",
    "los angeles rams": "Los Angeles Rams",
    "miami dolphins": "Miami Dolphins",
    "minnesota vikings": "Minnesota Vikings",
    "new england patriots": "New England Patriots",
    "new orleans saints": "New Orleans Saints",
    "new york giants": "New York Giants",
    "new york jets": "New York Jets",
    "philadelphia eagles": "Philadelphia Eagles",
    "pittsburgh steelers": "Pittsburgh Steelers",
    "san francisco 49ers": "San Francisco 49ers",
    "seattle seahawks": "Seattle Seahawks",
    "tampa bay buccaneers": "Tampa Bay Buccaneers",
    "tennessee titans": "Tennessee Titans",
    "washington commanders": "Washington Commanders",
    # Historical name changes
    "oakland raiders": "Las Vegas Raiders",
    "san diego chargers": "Los Angeles Chargers",
    "st louis rams": "Los Angeles Rams",
    "st. louis rams": "Los Angeles Rams",
    "washington redskins": "Washington Commanders",
    "washington football team": "Washington Commanders",
    # Australian source abbreviation variants
    "cardinals": "Arizona Cardinals",
    "falcons": "Atlanta Falcons",
    "ravens": "Baltimore Ravens",
    "bills": "Buffalo Bills",
    "panthers": "Carolina Panthers",
    "bears": "Chicago Bears",
    "bengals": "Cincinnati Bengals",
    "browns": "Cleveland Browns",
    "cowboys": "Dallas Cowboys",
    "broncos": "Denver Broncos",
    "lions": "Detroit Lions",
    "packers": "Green Bay Packers",
    "texans": "Houston Texans",
    "colts": "Indianapolis Colts",
    "jaguars": "Jacksonville Jaguars",
    "chiefs": "Kansas City Chiefs",
    "raiders": "Las Vegas Raiders",
    "chargers": "Los Angeles Chargers",
    "rams": "Los Angeles Rams",
    "dolphins": "Miami Dolphins",
    "vikings": "Minnesota Vikings",
    "patriots": "New England Patriots",
    "saints": "New Orleans Saints",
    "giants": "New York Giants",
    "jets": "New York Jets",
    "eagles": "Philadelphia Eagles",
    "steelers": "Pittsburgh Steelers",
    "49ers": "San Francisco 49ers",
    "seahawks": "Seattle Seahawks",
    "buccaneers": "Tampa Bay Buccaneers",
    "titans": "Tennessee Titans",
    "commanders": "Washington Commanders",
}

MLB_TEAM_ALIASES = {
    "arizona diamondbacks": "Arizona Diamondbacks",
    "atlanta braves": "Atlanta Braves",
    "baltimore orioles": "Baltimore Orioles",
    "boston red sox": "Boston Red Sox",
    "chicago cubs": "Chicago Cubs",
    "chicago white sox": "Chicago White Sox",
    "cincinnati reds": "Cincinnati Reds",
    "cleveland guardians": "Cleveland Guardians",
    "colorado rockies": "Colorado Rockies",
    "detroit tigers": "Detroit Tigers",
    "houston astros": "Houston Astros",
    "kansas city royals": "Kansas City Royals",
    "los angeles angels": "Los Angeles Angels",
    "los angeles dodgers": "Los Angeles Dodgers",
    "miami marlins": "Miami Marlins",
    "milwaukee brewers": "Milwaukee Brewers",
    "minnesota twins": "Minnesota Twins",
    "new york mets": "New York Mets",
    "new york yankees": "New York Yankees",
    "oakland athletics": "Oakland Athletics",
    "philadelphia phillies": "Philadelphia Phillies",
    "pittsburgh pirates": "Pittsburgh Pirates",
    "san diego padres": "San Diego Padres",
    "san francisco giants": "San Francisco Giants",
    "seattle mariners": "Seattle Mariners",
    "st. louis cardinals": "St. Louis Cardinals",
    "st louis cardinals": "St. Louis Cardinals",
    "tampa bay rays": "Tampa Bay Rays",
    "texas rangers": "Texas Rangers",
    "toronto blue jays": "Toronto Blue Jays",
    "washington nationals": "Washington Nationals",
    # Historical name changes
    "cleveland indians": "Cleveland Guardians",
    "florida marlins": "Miami Marlins",
    "tampa bay devil rays": "Tampa Bay Rays",
    "los angeles angels of anaheim": "Los Angeles Angels",
    "anaheim angels": "Los Angeles Angels",
    "montreal expos": "Washington Nationals",
    # Short names
    "diamondbacks": "Arizona Diamondbacks",
    "d-backs": "Arizona Diamondbacks",
    "braves": "Atlanta Braves",
    "orioles": "Baltimore Orioles",
    "red sox": "Boston Red Sox",
    "cubs": "Chicago Cubs",
    "white sox": "Chicago White Sox",
    "reds": "Cincinnati Reds",
    "guardians": "Cleveland Guardians",
    "indians": "Cleveland Guardians",
    "rockies": "Colorado Rockies",
    "tigers": "Detroit Tigers",
    "astros": "Houston Astros",
    "royals": "Kansas City Royals",
    "angels": "Los Angeles Angels",
    "dodgers": "Los Angeles Dodgers",
    "marlins": "Miami Marlins",
    "brewers": "Milwaukee Brewers",
    "twins": "Minnesota Twins",
    "mets": "New York Mets",
    "yankees": "New York Yankees",
    "athletics": "Oakland Athletics",
    "a's": "Oakland Athletics",
    "phillies": "Philadelphia Phillies",
    "pirates": "Pittsburgh Pirates",
    "padres": "San Diego Padres",
    "mariners": "Seattle Mariners",
    "cardinals": "St. Louis Cardinals",
    "rays": "Tampa Bay Rays",
    "rangers": "Texas Rangers",
    "blue jays": "Toronto Blue Jays",
    "nationals": "Washington Nationals",
}


def _resolve_team(name, alias_map):
    """Map a source team name to ESPN display name via alias dict."""
    if not name:
        return None
    key = str(name).strip().lower()
    return alias_map.get(key)


def _decimal_to_american(decimal_odds):
    """Convert decimal odds (e.g. 1.91) to American format (e.g. -110).

    Returns None if input is invalid.
    """
    try:
        d = float(decimal_odds)
        if d <= 1.0:
            return None
        if d >= 2.0:
            return int(round((d - 1) * 100))
        else:
            return int(round(-100 / (d - 1)))
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Download helpers
# ---------------------------------------------------------------------------

def _download_excel(url):
    """Download an Excel file and return an openpyxl Workbook."""
    import openpyxl

    logger.info("Downloading %s", url)
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) CoastCapital/1.0"
    }
    resp = requests.get(url, headers=headers, timeout=120)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    return wb


def _find_column_indices(header_row, target_columns):
    """Given a header row (list of cell values), return a dict of column_name → index.

    Uses case-insensitive substring matching.
    """
    indices = {}
    headers_lower = [str(h).strip().lower() if h else "" for h in header_row]
    for target in target_columns:
        target_lower = target.lower()
        for i, h in enumerate(headers_lower):
            if target_lower in h or h in target_lower:
                indices[target] = i
                break
    return indices


# ---------------------------------------------------------------------------
# NFL odds parsing
# ---------------------------------------------------------------------------

def _parse_nfl_odds(wb, start_year=None, end_year=None):
    """Parse NFL odds from the aussportsbetting.com workbook.

    Returns list of dicts ready for DB upsert.
    """
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Find header row (first row with "Date" in it)
    header_idx = 0
    for i, row in enumerate(rows):
        row_str = " ".join(str(c).lower() for c in row if c)
        if "date" in row_str and "home" in row_str:
            header_idx = i
            break

    header = rows[header_idx]
    col_map = _find_column_indices(header, [
        "date", "home", "away", "home odds", "away odds",
        "home line", "over/under",
    ])

    # Also try alternate column names
    if "home" not in col_map:
        col_map.update(_find_column_indices(header, [
            "home team", "away team", "home line spread",
            "total", "over under",
        ]))

    logger.info("NFL column mapping: %s", col_map)

    records = []
    for row in rows[header_idx + 1:]:
        try:
            date_val = row[col_map.get("date", 0)]
            if date_val is None:
                continue

            # Parse date
            if isinstance(date_val, datetime):
                game_date = date_val.date()
            else:
                # Try common date formats
                for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
                    try:
                        game_date = datetime.strptime(str(date_val).strip(), fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    continue

            year = game_date.year
            if start_year and year < start_year:
                continue
            if end_year and year > end_year:
                continue

            home_raw = row[col_map.get("home", col_map.get("home team", 1))]
            away_raw = row[col_map.get("away", col_map.get("away team", 2))]

            home_team = _resolve_team(home_raw, NFL_TEAM_ALIASES)
            away_team = _resolve_team(away_raw, NFL_TEAM_ALIASES)

            if not home_team or not away_team:
                logger.debug(
                    "NFL odds: unresolved team(s) home=%r away=%r on %s",
                    home_raw, away_raw, game_date,
                )
                continue

            # Odds
            home_decimal = row[col_map.get("home odds", 8)] if "home odds" in col_map else None
            away_decimal = row[col_map.get("away odds", 9)] if "away odds" in col_map else None
            spread = row[col_map.get("home line", col_map.get("home line spread", 10))] if any(
                k in col_map for k in ("home line", "home line spread")
            ) else None
            total = row[col_map.get("over/under", col_map.get("total", col_map.get("over under", 11)))] if any(
                k in col_map for k in ("over/under", "total", "over under")
            ) else None

            ml_home = _decimal_to_american(home_decimal)
            ml_away = _decimal_to_american(away_decimal)

            # Need at least one piece of odds data
            if spread is None and ml_home is None and total is None:
                continue

            try:
                spread_val = float(spread) if spread is not None else None
            except (TypeError, ValueError):
                spread_val = None

            try:
                total_val = float(total) if total is not None else None
            except (TypeError, ValueError):
                total_val = None

            records.append({
                "game_date": game_date,
                "home_team": home_team,
                "away_team": away_team,
                "spread": spread_val,
                "moneyline_home": ml_home,
                "moneyline_away": ml_away,
                "total_line": total_val,
                "sportsbook": "AusSportsBetting",
            })
        except Exception as exc:
            logger.debug("NFL odds parse error on row: %s", exc)
            continue

    logger.info("Parsed %d NFL odds records", len(records))
    return records


# ---------------------------------------------------------------------------
# MLB odds parsing
# ---------------------------------------------------------------------------

def _parse_mlb_odds(wb, year):
    """Parse MLB odds from a sports-statistics.com workbook for a single year.

    Returns list of dicts ready for DB upsert.
    """
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Find header row
    header_idx = 0
    for i, row in enumerate(rows):
        row_str = " ".join(str(c).lower() for c in row if c)
        if "date" in row_str and ("home" in row_str or "team" in row_str):
            header_idx = i
            break

    header = rows[header_idx]
    col_map = _find_column_indices(header, [
        "date", "home", "away", "home team", "away team",
        "home money", "away money", "home close",
        "away close", "close", "open", "total",
        "over/under", "run line", "home ml", "away ml",
    ])

    logger.info("MLB %d column mapping: %s (headers: %s)", year, col_map, header)

    records = []
    for row in rows[header_idx + 1:]:
        try:
            date_val = row[col_map.get("date", 0)]
            if date_val is None:
                continue

            if isinstance(date_val, datetime):
                game_date = date_val.date()
            else:
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%d/%m/%Y"):
                    try:
                        game_date = datetime.strptime(str(date_val).strip(), fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    continue

            # Home/away team
            home_col = col_map.get("home", col_map.get("home team"))
            away_col = col_map.get("away", col_map.get("away team"))
            if home_col is None or away_col is None:
                continue

            home_raw = row[home_col]
            away_raw = row[away_col]

            home_team = _resolve_team(home_raw, MLB_TEAM_ALIASES)
            away_team = _resolve_team(away_raw, MLB_TEAM_ALIASES)

            if not home_team or not away_team:
                logger.debug(
                    "MLB odds: unresolved team(s) home=%r away=%r on %s",
                    home_raw, away_raw, game_date,
                )
                continue

            # Moneylines — try various column names
            ml_home = None
            ml_away = None
            for key in ("home money", "home close", "home ml"):
                if key in col_map:
                    try:
                        ml_home = int(float(row[col_map[key]]))
                    except (TypeError, ValueError):
                        pass
                    break
            for key in ("away money", "away close", "away ml"):
                if key in col_map:
                    try:
                        ml_away = int(float(row[col_map[key]]))
                    except (TypeError, ValueError):
                        pass
                    break

            # Total/over-under
            total_val = None
            for key in ("total", "over/under", "close"):
                if key in col_map:
                    try:
                        total_val = float(row[col_map[key]])
                    except (TypeError, ValueError):
                        pass
                    break

            # Run line (MLB spread)
            spread_val = None
            if "run line" in col_map:
                try:
                    spread_val = float(row[col_map["run line"]])
                except (TypeError, ValueError):
                    pass

            if ml_home is None and ml_away is None and total_val is None:
                continue

            records.append({
                "game_date": game_date,
                "home_team": home_team,
                "away_team": away_team,
                "spread": spread_val,
                "moneyline_home": ml_home,
                "moneyline_away": ml_away,
                "total_line": total_val,
                "sportsbook": "SportsStatistics",
            })
        except Exception as exc:
            logger.debug("MLB odds parse error on row: %s", exc)
            continue

    logger.info("Parsed %d MLB odds records for %d", len(records), year)
    return records


# ---------------------------------------------------------------------------
# DB matching and upsert
# ---------------------------------------------------------------------------

def _match_and_upsert(schema, records):
    """Match parsed odds records to fact_game_results by (date, home, away)
    and insert into fact_market_odds.

    Returns dict with match/skip/insert counts.
    """
    if not records:
        return {"matched": 0, "skipped": 0, "inserted": 0}

    conn = get_connection(schema)
    try:
        cursor = conn.cursor()

        # Build a lookup of (game_date, home_team, away_team) → game_id
        cursor.execute(
            "SELECT game_id, DATE(game_date) AS gd, home_team, away_team "
            "FROM fact_game_results"
        )
        game_lookup = {}
        for game_id, gd, home, away in cursor.fetchall():
            key = (str(gd), home, away)
            game_lookup[key] = game_id

        logger.info("Loaded %d games from %s.fact_game_results for matching", len(game_lookup), schema)

        # Check existing odds to avoid duplicates
        cursor.execute(
            "SELECT game_id, sportsbook FROM fact_market_odds"
        )
        existing_odds = set()
        for gid, sb in cursor.fetchall():
            existing_odds.add((gid, sb))

        matched = 0
        skipped = 0
        inserted = 0

        for rec in records:
            key = (str(rec["game_date"]), rec["home_team"], rec["away_team"])
            game_id = game_lookup.get(key)

            if not game_id:
                skipped += 1
                continue

            matched += 1

            # Skip if we already have odds for this game+sportsbook
            if (game_id, rec["sportsbook"]) in existing_odds:
                continue

            cursor.execute(
                """
                INSERT INTO fact_market_odds
                    (game_id, sportsbook, spread, moneyline_home, moneyline_away,
                     total_line, market_timestamp)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    game_id,
                    rec["sportsbook"],
                    rec["spread"],
                    rec["moneyline_home"],
                    rec["moneyline_away"],
                    rec["total_line"],
                    rec["game_date"],
                ),
            )
            inserted += 1

        conn.commit()
        cursor.close()
        logger.info(
            "%s odds upsert: matched=%d, skipped=%d, inserted=%d",
            schema, matched, skipped, inserted,
        )
        return {"matched": matched, "skipped": skipped, "inserted": inserted}
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def backfill_historical_odds(sport, start_year=None, end_year=None):
    """Download historical odds and upsert into fact_market_odds.

    Parameters
    ----------
    sport : str
        ``"nfl"`` or ``"mlb"``.
    start_year : int, optional
        Filter to games on or after this year.
    end_year : int, optional
        Filter to games on or before this year.

    Returns
    -------
    dict with status, counts, and any unmatched team names.
    """
    sport = sport.lower()
    logger.info(
        "backfill_historical_odds: sport=%s start_year=%s end_year=%s",
        sport, start_year, end_year,
    )

    if sport == "nfl":
        wb = _download_excel(NFL_URL)
        records = _parse_nfl_odds(wb, start_year, end_year)
        wb.close()
        result = _match_and_upsert("nfl_silver", records)
        result["source"] = NFL_URL
        result["sport"] = "nfl"
        result["total_parsed"] = len(records)

    elif sport == "mlb":
        all_records = []
        sy = start_year or 2010
        ey = end_year or 2021
        for year in range(max(sy, 2010), min(ey, 2021) + 1):
            url = MLB_URL_TEMPLATE.format(year=year)
            try:
                wb = _download_excel(url)
                records = _parse_mlb_odds(wb, year)
                wb.close()
                all_records.extend(records)
            except Exception as exc:
                logger.warning("Failed to download MLB odds for %d: %s", year, exc)

        result = _match_and_upsert("mlb_silver", all_records)
        result["source"] = "sports-statistics.com"
        result["sport"] = "mlb"
        result["total_parsed"] = len(all_records)
        result["years_loaded"] = f"{sy}-{ey}"

    else:
        return {"status": "error", "error": f"Unsupported sport: {sport}. Use 'nfl' or 'mlb'."}

    result["status"] = "ok"
    return result
